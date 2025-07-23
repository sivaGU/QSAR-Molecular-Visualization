import streamlit as st
import os
from pathlib import Path
import base64
import subprocess
import sys
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np

# Page configuration
st.set_page_config(
    page_title="QSAR Molecular Visualization Tool",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
    body {
        background: #f4f6fa;
    }
    .main-header {
        font-size: 2.8rem;
        font-weight: 900;
        color: #2563eb;
        text-align: center;
        margin-bottom: 0.2rem;
        letter-spacing: -1px;
    }
    .sub-header {
        font-size: 1.25rem;
        color: #4b5563;
        text-align: center;
        margin-bottom: 2.5rem;
        font-weight: 400;
    }
    .receptor-card {
        background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
        padding: 2.2rem 2rem 2rem 2rem;
        border-radius: 18px;
        color: white;
        text-align: center;
        box-shadow: 0 6px 24px rgba(239,68,68,0.08);
        margin-bottom: 1.5rem;
    }
    .receptor-card-beta {
        background: linear-gradient(135deg, #14b8a6 0%, #0d9488 100%);
        padding: 2.2rem 2rem 2rem 2rem;
        border-radius: 18px;
        color: white;
        text-align: center;
        box-shadow: 0 6px 24px rgba(20,184,166,0.08);
        margin-bottom: 1.5rem;
    }
    .info-block {
        background: rgba(255, 255, 255, 0.18);
        padding: 1.1rem 1.5rem;
        border-radius: 12px;
        margin: 1.2rem 0.5rem 1.2rem 0;
        display: inline-block;
        font-size: 1.1rem;
        font-weight: 600;
        box-shadow: 0 2px 8px rgba(0,0,0,0.03);
    }
    .stButton > button {
        background: linear-gradient(90deg, #2563eb 0%, #1e40af 100%);
        color: white;
        border: none;
        border-radius: 25px;
        padding: 0.7rem 2.2rem;
        font-weight: bold;
        font-size: 1.1rem;
        margin-top: 0.7rem;
        margin-bottom: 0.7rem;
        box-shadow: 0 2px 8px rgba(37,99,235,0.08);
        transition: all 0.2s;
    }
    .stButton > button:hover {
        background: linear-gradient(90deg, #1e40af 0%, #2563eb 100%);
        transform: translateY(-2px) scale(1.04);
    }
    .viewer-container {
        border: 2px solid #e5e7eb;
        border-radius: 16px;
        padding: 18px 18px 0 18px;
        margin: 32px 0 24px 0;
        background: #f9fafb;
        box-shadow: 0 4px 16px rgba(0,0,0,0.04);
        max-width: 900px;
        margin-left: auto;
        margin-right: auto;
    }
    .ngl-title {
        font-size: 1.25rem;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 0.5rem;
        letter-spacing: -0.5px;
    }
    .stSelectbox > div {
        font-size: 1.1rem;
        font-weight: 500;
    }
    .stSidebar {
        background: #f1f5f9 !important;
    }
    .stSidebar [data-testid="stSidebarNav"] {
        margin-top: 2rem;
    }
    .stSidebar h1, .stSidebar h2, .stSidebar h3, .stSidebar h4 {
        color: #2563eb !important;
    }
    .st-expanderHeader {
        font-size: 1.1rem;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

def get_ligand_list(folder_name):
    folder_path = Path(folder_name)
    if not folder_path.exists():
        return []
    ligands = []
    
    # Handle different file naming patterns
    if "T50" in folder_name:
        # T50 files: either *_top_complex.pdb or *_out_complex.pdb
        for file in folder_path.glob("*_complex.pdb"):
            ligand_name = file.stem.replace("_top_complex", "").replace("_out_complex", "")
            ligands.append(ligand_name)
    else:
        # CE files: combined_*_out.pdb
        for file in folder_path.glob("combined_*.pdb"):
            ligand_name = file.stem.replace("combined_", "").replace("_out", "")
            ligands.append(ligand_name)
    
    return sorted(ligands)

def create_ngl_viewer(pdb_content, structure_name):
    pdb_encoded = base64.b64encode(pdb_content.encode()).decode()
    html_code = f"""
    <div class='viewer-container'>
        <div id='ngl-viewer' style='width: 100%; height: 520px; border: 1px solid #ddd; border-radius: 12px;'></div>
    </div>
    <script src='https://unpkg.com/ngl@0.10.4/dist/ngl.js'></script>
    <script>
        var stage = new NGL.Stage("ngl-viewer");
        stage.setParameters({{ backgroundColor: "white" }});
        
        var pdbData = atob("{pdb_encoded}");
        stage.loadFile(new Blob([pdbData], {{type: "chemical/x-pdb"}}), {{ext: "pdb"}}).then(function (component) {{
            // Default representation - let NGL Viewer decide based on PDB content
            component.addRepresentation("cartoon");
            
            // Try different selections for ligands
            component.addRepresentation("ball+stick", {{ sele: "hetero" }});
            component.addRepresentation("ball+stick", {{ sele: "UNL" }});
            component.addRepresentation("ball+stick", {{ sele: "not protein" }});
            
            component.autoView();
        }});
        
        // Prevent page scroll when zooming
        var viewerDiv = document.getElementById("ngl-viewer");
        viewerDiv.addEventListener('wheel', function(event) {{
            event.preventDefault();
        }}, {{ passive: false }});
    </script>
    """
    return html_code

def open_pdb_file(file_path):
    try:
        if sys.platform == "win32":
            os.startfile(str(file_path))
        elif sys.platform == "darwin":
            subprocess.run(["open", str(file_path)])
        else:
            subprocess.run(["xdg-open", str(file_path)])
        return True
    except Exception as e:
        st.error(f"Could not open the PDB file. Error: {str(e)}")
        return False

def show_ce_ligand_comparison():
    import openpyxl
    import pandas as pd
    import plotly.graph_objects as go
    st.markdown("## CE Ligand Comparison: Alpha vs Beta Docking Scores")
    st.markdown("Compare the docking scores for each commonly exposed ligand between ERα and ERβ.")

    # Load Table S5 (Commonly Exposed) with correct relative path
    wb = openpyxl.load_workbook(r'C:\Users\madas\Downloads\Final GitHub Submission\Supplementary Tables\Supplementary Table 5 (Table S5).xlsx')
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(header)}
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            casrn = row[idx['CASRN']]
            alpha = float(row[idx['Alpha Docking Score (kcal/mol)']])
            beta = float(row[idx['Beta Docking Score (kcal/mol)']])
            diff = alpha - beta
            data.append({
                'CASRN': casrn,
                'Alpha Docking Score': alpha,
                'Beta Docking Score': beta,
                'Difference (Alpha - Beta)': diff
            })
        except:
            pass
    df = pd.DataFrame(data)
    st.markdown("### 📋 Docking Score Comparison Table")
    st.dataframe(df, use_container_width=True)

    # Scatter plot: Alpha vs Beta docking score
    st.markdown("### 🎯 Scatter Plot: Alpha vs Beta Docking Score")
    fig_scatter = go.Figure()
    fig_scatter.add_trace(go.Scatter(
        x=df['Alpha Docking Score'],
        y=df['Beta Docking Score'],
        mode='markers',
        marker=dict(color='#2563eb', size=8, opacity=0.7),
        text=df['CASRN'],
        showlegend=False
    ))
    fig_scatter.add_trace(go.Scatter(
        x=[df['Alpha Docking Score'].min(), df['Alpha Docking Score'].max()],
        y=[df['Alpha Docking Score'].min(), df['Alpha Docking Score'].max()],
        mode='lines',
        line=dict(color='gray', dash='dash'),
        name='y=x (Equal Score)'
    ))
    fig_scatter.update_layout(
        title="Alpha vs Beta Docking Score (CE Ligands)",
        xaxis_title="Alpha Docking Score",
        yaxis_title="Beta Docking Score",
        height=400,
        xaxis=dict(range=[df['Alpha Docking Score'].min()-0.5, df['Alpha Docking Score'].max()+0.5], fixedrange=True),
        yaxis=dict(range=[df['Beta Docking Score'].min()-0.5, df['Beta Docking Score'].max()+0.5], fixedrange=True),
        dragmode=False
    )
    st.plotly_chart(fig_scatter, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})

    # Histogram of differences
    st.markdown("### 🧬 Histogram: Alpha - Beta Docking Score Differences")
    fig_hist = go.Figure()
    fig_hist.add_trace(go.Histogram(
        x=df['Difference (Alpha - Beta)'],
        marker_color="#14b8a6",
        nbinsx=20,
        showlegend=False
    ))
    fig_hist.update_layout(
        title="Distribution of Docking Score Differences (Alpha - Beta)",
        xaxis_title="Alpha - Beta Docking Score",
        yaxis_title="Ligand Count",
        height=350,
        xaxis=dict(fixedrange=True),
        yaxis=dict(fixedrange=True),
        dragmode=False
    )
    st.plotly_chart(fig_hist, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})

    st.markdown("### 💡 Key Insights")
    st.info(f"""
    - Most CE ligands have similar docking scores for Alpha and Beta, but some show notable differences.
    - The scatter plot shows the correlation and outliers between Alpha and Beta docking.
    - The histogram shows the distribution of Alpha-Beta differences across all CE ligands.
    """)

def show_chemical_descriptor_analysis():
    import pandas as pd
    import plotly.graph_objects as go
    import plotly.express as px
    st.markdown("## Chemical Descriptor Analysis")
    st.markdown("Property distributions and trends based on directional coefficient across QSPR models.")

    # Descriptor coefficients from the provided table
    data = [
        ["# of H-Bond Acceptors", 0.29, 0.21, -0.52, -1.01],
        ["# of H-Bond Donors", 0.26, 0.22, 0.03, -0.02],
        ["LogD", 0.20, 0.40, 0.21, 0.17],
        ["LogP", -0.04, -0.28, -0.24, 0.30],
        ["Average Mass", -0.42, -0.23, -0.99, -0.67],
        ["Boiling Point", 1.55, 2.60, 0.23, 1.07],
        ["Density", 0.17, 0.33, 0.41, 0.15],
        ["Enthalpy of Vaporization", -1.56, -2.83, -0.27, -0.37],
        ["F+ Max", 0.03, 0.08, -0.08, -0.06],
        ["HOMO", -0.19, 0.07, 0.05, -0.04],
        ["LUMO", 0.20, 0.08, -0.06, 0.02],
        ["Polar Surface Area", -0.22, -0.20, 0.71, 0.58],
        ["Polarizability", 0.77, 1.04, 0.32, -1.21],
        ["Surface Tension", -0.39, -0.56, -0.78, -0.12],
        ["# of Freely Rotating Bonds", -0.44, -0.65, 0.96, 1.35],
    ]
    df = pd.DataFrame(data, columns=[
        "Descriptor",
        "Top 50 α",
        "Top 50 β",
        "CE α",
        "CE β"
    ])
    st.markdown("### 📋 QSPR Model Coefficients Table")
    st.dataframe(df, use_container_width=True)

    # Heatmap
    st.markdown("### 🔥 Descriptor Coefficient Heatmap")
    st.write("This heatmap shows the direction (red=negative, green=positive) and strength (color intensity) of each descriptor's effect in each model.")
    heatmap_df = df.set_index("Descriptor").T
    fig_heatmap = px.imshow(
        heatmap_df,
        color_continuous_scale=["#ef4444", "#f9fafb", "#22c55e"],
        aspect="auto",
        labels=dict(x="Descriptor", y="Model", color="Coefficient"),
        zmin=-3, zmax=3
    )
    fig_heatmap.update_layout(
        height=400,
        xaxis_title="Descriptor",
        yaxis_title="Model",
        coloraxis_colorbar=dict(title="Coefficient Value"),
        dragmode=False
    )
    st.plotly_chart(fig_heatmap, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})

    # Top Influencers
    st.markdown("### ⭐ Top Influential Descriptors per Model")
    st.write("For each model, the top 3 positive and top 3 negative coefficients are shown. This highlights the most important features for binding affinity prediction.")
    models = ["Top 50 α", "Top 50 β", "CE α", "CE β"]
    for model in models:
        st.markdown(f"#### {model} Model")
        top_pos = df.nlargest(3, model)[["Descriptor", model]]
        top_neg = df.nsmallest(3, model)[["Descriptor", model]]
        top = pd.concat([top_pos, top_neg])
        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(
            y=top["Descriptor"],
            x=top[model],
            orientation='h',
            marker_color=["#22c55e" if v >= 0 else "#ef4444" for v in top[model]],
            showlegend=False
        ))
        fig_bar.update_layout(
            title=f"Top Influential Descriptors ({model})",
            xaxis_title="Coefficient Value",
            yaxis_title="Descriptor",
            height=350,
            yaxis=dict(autorange="reversed", fixedrange=True),
            xaxis=dict(fixedrange=True),
            dragmode=False
        )
        st.plotly_chart(fig_bar, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})

    st.markdown("### 💡 Key Insights")
    st.info("""
    - The heatmap allows rapid comparison of descriptor effects across all models.
    - Top influencers highlight which features most strongly increase or decrease predicted binding affinity in each model.
    - Descriptor importance and directionality can shift between models, reflecting different binding mechanisms or ligand sets.
    """)

def show_qsar_results():
    import pandas as pd
    st.markdown("## QSAR Results: Large Set Model Performance")
    st.markdown("Explore the performance of large set QSAR models for ERα and ERβ, including the effect of outlier removal.")

    # Corrected R² summary table from images (no Receptor column)
    r2_data = [
        ["Alpha", "Original", 0.451],
        ["Alpha", "10% Outliers Removed", 0.625],
        ["Alpha", "20% Outliers Removed", 0.660],
        ["Beta", "Original", 0.493],
        ["Beta", "10% Outliers Removed", 0.638],
        ["Beta", "20% Outliers Removed", 0.682],
    ]
    df_r2 = pd.DataFrame(r2_data, columns=["Receptor", "Refinement Step", "R²"])
    st.markdown("### 📊 Model R² Summary Table")
    st.dataframe(df_r2, use_container_width=True)

    # Model improvement visualization
    st.markdown("### 📈 Model Improvement Analysis")
    st.markdown("Visual comparison of R² improvements across refinement steps for Alpha and Beta receptors.")
    
    import plotly.graph_objects as go
    
    # Create grouped bar chart
    fig = go.Figure()
    
    # Alpha data
    alpha_data = df_r2[df_r2['Receptor'] == 'Alpha']
    fig.add_trace(go.Bar(
        name='Alpha',
        x=alpha_data['Refinement Step'],
        y=alpha_data['R²'],
        marker_color='#3b82f6',
        text=alpha_data['R²'].round(3),
        textposition='auto',
    ))
    
    # Beta data
    beta_data = df_r2[df_r2['Receptor'] == 'Beta']
    fig.add_trace(go.Bar(
        name='Beta',
        x=beta_data['Refinement Step'],
        y=beta_data['R²'],
        marker_color='#ef4444',
        text=beta_data['R²'].round(3),
        textposition='auto',
    ))
    
    fig.update_layout(
        title="R² Improvement Across Refinement Steps",
        xaxis_title="Refinement Step",
        yaxis_title="R² Value",
        barmode='group',
        height=500,
        showlegend=True,
        yaxis=dict(range=[0, 0.8]),
        plot_bgcolor='white',
        paper_bgcolor='white',
    )
    
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})
    
    # Improvement summary
    st.markdown("#### Key Insights:")
    col1, col2 = st.columns(2)
    
    with col1:
        alpha_improvement = ((alpha_data.iloc[2]['R²'] - alpha_data.iloc[0]['R²']) / alpha_data.iloc[0]['R²'] * 100).round(1)
        st.metric("Alpha Improvement", f"{alpha_improvement}%", f"0.451 → 0.660")
    
    with col2:
        beta_improvement = ((beta_data.iloc[2]['R²'] - beta_data.iloc[0]['R²']) / beta_data.iloc[0]['R²'] * 100).round(1)
        st.metric("Beta Improvement", f"{beta_improvement}%", f"0.493 → 0.682")

def main():
    st.markdown('<h1 class="main-header">🧬 QSAR Molecular Visualization Tool</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Interactive 3D visualization of ERα and ERβ receptor-PFAS ligand structures</p>', unsafe_allow_html=True)
    st.sidebar.title("Navigation")
    
    # Initialize session state
    if 'page' not in st.session_state:
        st.session_state.page = "Home"
    
    # Handle page navigation from buttons
    if st.session_state.page != "Home":
        page = st.session_state.page
        st.session_state.page = "Home"  # Reset for next time
    else:
        page = st.sidebar.selectbox(
            "Choose a page:",
            [
                "Home",
                "ERα Receptor",
                "ERβ Receptor",
                "Data Analysis Dashboard",
                "CE Ligand Comparison",
                "Chemical Descriptor Analysis",
                "QSAR Results",
                "About"
            ]
        )
    
    if page == "Home":
        show_home_page()
    elif page == "ERα Receptor":
        show_alpha_page()
    elif page == "ERβ Receptor":
        show_beta_page()
    elif page == "Data Analysis Dashboard":
        show_data_analysis_dashboard()
    elif page == "CE Ligand Comparison":
        show_ce_ligand_comparison()
    elif page == "Chemical Descriptor Analysis":
        show_chemical_descriptor_analysis()
    elif page == "QSAR Results":
        show_qsar_results()
    elif page == "About":
        show_about_page()

def show_home_page():
    st.markdown("## Welcome to QSAR Molecular Visualization Tool")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
        <div class="receptor-card">
            <h2 style='font-size:2rem;font-weight:800;'>🧬 ERα Receptor</h2>
            <p style='font-size:1.1rem;'>Estrogen Receptor Alpha - Primary target for estrogen signaling</p>
            <div class="info-block">
                <strong>106</strong><br>
                <small>Ligands</small>
            </div>
            <div class="info-block">
                <strong>3D</strong><br>
                <small>Visualization</small>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown("""
        <div class="receptor-card-beta">
            <h2 style='font-size:2rem;font-weight:800;'>🧬 ERβ Receptor</h2>
            <p style='font-size:1.1rem;'>Estrogen Receptor Beta - Secondary estrogen receptor subtype</p>
            <div class="info-block">
                <strong>106</strong><br>
                <small>Ligands</small>
            </div>
            <div class="info-block">
                <strong>3D</strong><br>
                <small>Visualization</small>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("""
    ### How to Use This Tool
    1. **Select a Receptor**: Choose between ERα (primary estrogen receptor) or ERβ (secondary estrogen receptor)
    2. **Choose a Ligand**: Select from 106 available PFAS ligands
    3. **Visualize**: Use the embedded 3D viewer or download the PDB file
    ### Available Features
    - **PFAS Ligands**: Each ligand is combined with the selected receptor
    - **3D Visualization**: Interactive molecular viewer built into the browser
    - **Cross-platform**: Works on any device with a web browser
    - **Easy Download**: Direct download links for all combined structures
    """)

def show_alpha_page():
    st.markdown("## 🧬 ERα Receptor Visualization")
    st.markdown("**Estrogen Receptor Alpha - Primary target for estrogen signaling**")
    
    # Select dataset
    dataset = st.selectbox(
        "Choose a dataset:",
        ["Commonly Exposed Set", "Top 50 Set"],
        index=0
    )
    
    # Get ligands based on selected dataset
    if dataset == "Commonly Exposed Set":
        folder_name = "Alpha_CE_Combined"
    else:  # Top 50 Set
        folder_name = "Alpha_T50_Combined"
    
    alpha_ligands = get_ligand_list(folder_name)
    if not alpha_ligands:
        st.error(f"No combined PDB files found in '{folder_name}' folder. Please run the combine_pdb.py script first.")
        return
    
    st.markdown("### Select a Ligand")
    selected_ligand = st.selectbox(
        f"Choose a PFAS ligand to visualize with ERα ({dataset}):",
        alpha_ligands,
        index=0
    )
    if selected_ligand:
        # Different file naming patterns for different datasets
        if dataset == "Top 50 Set":
            file_name = f"{selected_ligand}_top_complex.pdb"
        else:  # Commonly Exposed Set
            file_name = f"combined_{selected_ligand}_out.pdb"
        
        file_path = Path(folder_name) / file_name
        if file_path.exists():
            file_size = file_path.stat().st_size / 1024
            st.info(f"**File Size:** {file_size:.1f} KB")
            pdb_content = file_path.read_text()
            st.download_button(
                label="📁 Download PDB File",
                data=pdb_content,
                file_name=file_name,
                mime="chemical/x-pdb",
                key="alpha_download"
            )
            st.markdown("### 🧬 Interactive 3D Molecular Viewer")
            st.markdown("**Rotate, zoom, and explore the molecular structure directly in your browser**")
            viewer_html = create_ngl_viewer(pdb_content, f"ERα + {selected_ligand}")
            st.components.v1.html(viewer_html, height=600)
            st.markdown("""
            **Viewer Controls:**
            - **Mouse**: Rotate the structure
            - **Scroll**: Zoom in/out (page will not scroll)
            - **Right-click + drag**: Pan the view
            - **Double-click**: Reset view
            """)
            st.markdown("### File Preview")
            with st.expander("View PDB file content (first 50 lines)"):
                lines = pdb_content.split('\n')[:50]
                st.code('\n'.join(lines))
        else:
            st.error(f"❌ Combined PDB file not found: {file_path}")

def show_beta_page():
    st.markdown("## 🧬 ERβ Receptor Visualization")
    st.markdown("**Estrogen Receptor Beta - Secondary estrogen receptor subtype**")
    
    # Select dataset
    dataset = st.selectbox(
        "Choose a dataset:",
        ["Commonly Exposed Set", "Top 50 Set"],
        index=0
    )
    
    # Get ligands based on selected dataset
    if dataset == "Commonly Exposed Set":
        folder_name = "Beta_CE_Combined"
    else:  # Top 50 Set
        folder_name = "Beta_T50_Combined"
    
    beta_ligands = get_ligand_list(folder_name)
    if not beta_ligands:
        st.error(f"No combined PDB files found in '{folder_name}' folder. Please run the combine_pdb.py script first.")
        return
    
    st.markdown("### Select a Ligand")
    selected_ligand = st.selectbox(
        f"Choose a PFAS ligand to visualize with ERβ ({dataset}):",
        beta_ligands,
        index=0
    )
    if selected_ligand:
        # Different file naming patterns for different datasets
        if dataset == "Top 50 Set":
            file_name = f"{selected_ligand}_out_complex.pdb"
        else:  # Commonly Exposed Set
            file_name = f"combined_{selected_ligand}_out.pdb"
        
        file_path = Path(folder_name) / file_name
        if file_path.exists():
            file_size = file_path.stat().st_size / 1024
            st.info(f"**File Size:** {file_size:.1f} KB")
            pdb_content = file_path.read_text()
            st.download_button(
                label="📁 Download PDB File",
                data=pdb_content,
                file_name=file_name,
                mime="chemical/x-pdb",
                key="beta_download"
            )
            st.markdown("### 🧬 Interactive 3D Molecular Viewer")
            st.markdown("**Rotate, zoom, and explore the molecular structure directly in your browser**")
            viewer_html = create_ngl_viewer(pdb_content, f"ERβ + {selected_ligand}")
            st.components.v1.html(viewer_html, height=600)
            st.markdown("""
            **Viewer Controls:**
            - **Mouse**: Rotate the structure
            - **Scroll**: Zoom in/out (page will not scroll)
            - **Right-click + drag**: Pan the view
            - **Double-click**: Reset view
            """)
            st.markdown("### File Preview")
            with st.expander("View PDB file content (first 50 lines)"):
                lines = pdb_content.split('\n')[:50]
                st.code('\n'.join(lines))
        else:
            st.error(f"❌ Combined PDB file not found: {file_path}")

def show_data_analysis_dashboard():
    st.markdown("## 📊 Data Analysis Dashboard")
    st.markdown("**Statistical summaries and visualizations of the 4 datasets**")

    # Real statistics from parsed Excel files (means only)
    summary_data = [
        {
            "Dataset": "CE Ligands",
            "Ligand Count": 55,
            "Alpha Docking Score": "-8.46",
            "Beta Docking Score": "-8.56",
            "LogP": "8.89",
            "MW": "567.95",
            "PSA": "40.02"
        },
        {
            "Dataset": "Alpha T50",
            "Ligand Count": 51,
            "Alpha Docking Score": "-11.27",
            "Beta Docking Score": "—",
            "LogP": "7.71",
            "MW": "507.02",
            "PSA": "17.08"
        },
        {
            "Dataset": "Beta T50",
            "Ligand Count": 51,
            "Alpha Docking Score": "—",
            "Beta Docking Score": "-11.25",
            "LogP": "5.76",
            "MW": "555.94",
            "PSA": "12.04"
        }
    ]
    df_summary = pd.DataFrame(summary_data)
    st.markdown("### 📋 Dataset Comparison")
    st.dataframe(df_summary, use_container_width=True)

    # Prepare data for charts (means only)
    chart_data = pd.DataFrame({
        "Dataset": ["CE Ligands (Alpha)", "CE Ligands (Beta)", "Alpha T50", "Beta T50"],
        "Docking Score Mean": [-8.46, -8.56, -11.27, -11.25],
        "Type": ["CE", "CE", "T50", "T50"],
        "Receptor": ["Alpha", "Beta", "Alpha", "Beta"]
    })

    # Chart: Docking Score Means (no error bars)
    st.markdown("### 🎯 Docking Score Comparison")
    fig_docking = go.Figure()
    colors = ["#2563eb", "#14b8a6", "#f59e42", "#e11d48"]
    for i, row in chart_data.iterrows():
        fig_docking.add_trace(go.Bar(
            x=[row["Dataset"]],
            y=[row["Docking Score Mean"]],
            name=row["Dataset"],
            marker_color=colors[i]
        ))
    min_ds = min(chart_data["Docking Score Mean"])
    max_ds = max(chart_data["Docking Score Mean"])
    fig_docking.update_layout(
        barmode='group',
        title="Average Docking Scores",
        yaxis_title="Docking Score (kcal/mol)",
        xaxis_title="Dataset",
        height=400,
        yaxis=dict(range=[min_ds-0.5, max_ds+0.5], fixedrange=True),
        dragmode=False
    )
    st.plotly_chart(fig_docking, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})

    # Chart: Descriptor Comparison (means only)
    st.markdown("### 🧬 Descriptor Comparison")
    desc_data = pd.DataFrame({
        "Dataset": ["CE Ligands", "Alpha T50", "Beta T50"],
        "LogP": [8.89, 7.71, 5.76],
        "MW": [567.95, 507.02, 555.94],
        "PSA": [40.02, 17.08, 12.04]
    })
    # LogP
    fig_logp = go.Figure()
    for i, row in desc_data.iterrows():
        fig_logp.add_trace(go.Bar(
            x=[row["Dataset"]],
            y=[row["LogP"]],
            name=row["Dataset"],
            marker_color=colors[i]
        ))
    min_logp = min(desc_data["LogP"])
    max_logp = max(desc_data["LogP"])
    fig_logp.update_layout(
        barmode='group',
        title="Average LogP",
        yaxis_title="LogP",
        xaxis_title="Dataset",
        height=350,
        yaxis=dict(range=[min_logp-0.5, max_logp+0.5], fixedrange=True),
        dragmode=False
    )
    st.plotly_chart(fig_logp, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})
    # MW
    fig_mw = go.Figure()
    for i, row in desc_data.iterrows():
        fig_mw.add_trace(go.Bar(
            x=[row["Dataset"]],
            y=[row["MW"]],
            name=row["Dataset"],
            marker_color=colors[i]
        ))
    min_mw = min(desc_data["MW"])
    max_mw = max(desc_data["MW"])
    fig_mw.update_layout(
        barmode='group',
        title="Average Molecular Weight",
        yaxis_title="Molecular Weight (g/mol)",
        xaxis_title="Dataset",
        height=350,
        yaxis=dict(range=[min_mw-10, max_mw+10], fixedrange=True),
        dragmode=False
    )
    st.plotly_chart(fig_mw, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})
    # PSA
    fig_psa = go.Figure()
    for i, row in desc_data.iterrows():
        fig_psa.add_trace(go.Bar(
            x=[row["Dataset"]],
            y=[row["PSA"]],
            name=row["Dataset"],
            marker_color=colors[i]
        ))
    min_psa = min(desc_data["PSA"])
    max_psa = max(desc_data["PSA"])
    fig_psa.update_layout(
        barmode='group',
        title="Average Polar Surface Area",
        yaxis_title="Polar Surface Area",
        xaxis_title="Dataset",
        height=350,
        yaxis=dict(range=[min_psa-2, max_psa+2], fixedrange=True),
        dragmode=False
    )
    st.plotly_chart(fig_psa, use_container_width=True, config={"displayModeBar": False, "staticPlot": True})

    st.markdown("### 💡 Key Insights")
    st.info("""
    - **CE Ligands** have higher LogP, MW, and PSA than T50 sets, but lower docking scores (weaker binding).
    - **Alpha T50 and Beta T50** have similar strong docking scores, but their ligand properties differ.
    - **Beta T50** ligands have lower LogP and PSA but higher MW than Alpha T50.
    """)

def show_about_page():
    st.markdown("## About QSAR Molecular Visualization Tool")
    st.markdown("""
    ### Overview
    This tool provides an interactive web-based interface for visualizing QSAR (Quantitative Structure-Activity Relationship) molecular structures, specifically focusing on Estrogen Receptor (ER) interactions with PFAS ligands.
    ### Features
    - **Dual Receptor Support**: ERα and ERβ receptor visualization
    - **PFAS Ligands**: Comprehensive library of per- and polyfluoroalkyl substances
    - **Combined Structures**: Pre-combined receptor-ligand complexes
    - **Embedded 3D Viewer**: Interactive molecular visualization using NGL Viewer
    - **Multiple Output Options**: Download, open with default viewer, or copy file paths
    - **Cross-platform Compatibility**: Works on any device with a web browser
    - **No Installation Required**: Everything works in your browser
    ### Technical Details
    - **File Format**: PDB (Protein Data Bank) format
    - **Combined Files**: Each file contains both receptor and ligand structures
    - **3D Viewer**: NGL Viewer for interactive molecular visualization
    - **File Organization**: 
      - `Alpha Combined/`: ERα receptor + ligand complexes
      - `Beta Combined/`: ERβ receptor + ligand complexes
    ### Viewer Features
    - **Interactive 3D Visualization**: Rotate, zoom, and pan molecular structures
    - **Multiple Representations**: Cartoon and ball+stick views
    - **Color Coding**: Chain-based coloring for easy identification
    - **Hetero Atoms**: Ligands displayed as ball+stick representation
    - **Responsive Design**: Works on desktop, tablet, and mobile devices
    ### Recommended Molecular Viewers (for downloaded files)
    - **PyMOL**: Professional molecular visualization
    - **VMD**: Visual Molecular Dynamics
    - **ChimeraX**: UCSF ChimeraX
    - **Jmol**: Java-based molecular viewer
    - **Online viewers**: NGL Viewer, Mol* Viewer
    ### Usage Instructions
    1. Navigate to the desired receptor page (ERα or ERβ)
    2. Select a ligand from the dropdown menu
    3. Use the embedded 3D viewer to explore the structure
    4. Choose additional actions:
       - Download the PDB file
       - Open with your default molecular viewer
       - Copy the file path for manual access
    ### Data Source
    The combined PDB files are generated from individual receptor and ligand structures, providing ready-to-use complexes for molecular visualization and analysis.
    """)

if __name__ == "__main__":
    main() 